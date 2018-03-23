# -*- coding: utf-8 -*-
import argparse
import logging
import os
import sys
from argparse import RawTextHelpFormatter

from openpyxl import load_workbook

import pyForestLASTools


class UpdateLASJulianDate(object):
    @staticmethod
    def Header():
        print('Update LAS Julian date  v0.9')
        print

    @staticmethod
    def ParseCmdLine():
        # updateLASJulianDate.py G:\TRANSECTS\LAZ\METADATA\t-0002.xlsx -p G:\TRANSECTS\LAZ\METADATA
        parser = argparse.ArgumentParser(description='Process python scripts in multiprocessing mode.',
                                         formatter_class=RawTextHelpFormatter)
        parser.add_argument('lasfile', help='LAS/LAZ file to be updated.')
        parser.add_argument('-p', '--xlsxpath', help='Path to xlsx files.', type=str, default='')
        parser.add_argument('-d', '--checkdate', help='Check date field.', action='store_true', default=False)
        parser.add_argument("-v", "--verbose", type=int, help="Show intermediate messages.", default=0)
        parser.add_argument("-l", "--log", type=str, default=None, help="Logs to a file. Default 'None'.")
        try:
            return parser.parse_args()
        except:
            print
            sys.exc_info()[0]
            raise

    def __init__(self, lasfile, xlsxpath, checkdate, verbose):
        self.lasfile = lasfile
        self.xlsxpath = xlsxpath
        self.checkdate = checkdate
        self.verbose = verbose
        self.pyFLT = pyForestLASTools.pyForestLASTools(lasfile)

    def UpdateJulianDate(self, julianDay, year):
        result, error = self.pyFLT.updateLASFile(self.lasfile, '', '-set_file_creation {0} {1}'.format(julianDay, year))

    def Process(self):
        xlsxfname = os.path.join(self.xlsxpath,
                                 os.path.splitext(os.path.basename(self.lasfile.replace('NP_', '')))[0]) + '.xlsx'
        if not os.path.isfile(xlsxfname):
            msg = 'File not found: {0}.'.format(xlsxfname)
            logging.error(msg)
            raise Exception(msg)

        wb = load_workbook(xlsxfname)
        ws = wb.active
        if ws['B5'].value.startswith('Data(s) de aquisi'):
            acquisitionDate = ws['C5'].value
            try:
                julianDay = acquisitionDate.timetuple()
            except:
                return False, acquisitionDate
            if self.checkdate:
                return True, ''
            result, error = self.UpdateJulianDate(julianDay.tm_yday, julianDay.tm_year)
            logging.info("File {0} updated.".format(xlsxfname))

            return True, result
        else:
            msg = 'Invalid update date in {0}.'.format(xlsxfname)
            logging.error(msg)
            raise Exception(msg)


if __name__ == "__main__":
    reload(sys)
    sys.setdefaultencoding('utf-8')
    UpdateLASJulianDate.Header()
    try:
        args = UpdateLASJulianDate.ParseCmdLine()
        if args.log:
            logging.basicConfig(filename=args.log, level=logging.INFO)
            logging.getLogger().addHandler(logging.StreamHandler())
        else:
            logging.basicConfig(level=logging.INFO)
        uljd = UpdateLASJulianDate(args.lasfile, args.xlsxpath, args.checkdate, args.verbose)
        result, date = uljd.Process()
        if not result:
            print >> sys.stderr, 'Error in {0}: Flight date {1} invalid.'.format(args.lasfile, date)
        else:
            if (isinstance(date, basestring)) and ((date.upper().find("ERROR") > 0)):
                print >> sys.stderr, date

    except Exception as e:
        raise RuntimeError("Unexpected error: {}".format(e))
